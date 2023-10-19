from threading import Thread
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
import re
import time
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.formula.translate import Translator
from openpyxl.styles import PatternFill, Font, Border
from openpyxl.formula.translate import Translator
from openpyxl.utils import quote_sheetname
from openpyxl.worksheet.datavalidation import DataValidation


from types import SimpleNamespace
import pandas as pd


class StyleTemplatePandas:
    def __init__(
        self,
        file_template_path: str,
        save_path: str,
        file_path=False,
        adictional_rules_to_apply:dict=False,
        datas_validation_to_apply:dict=False,
        show_gridlines:bool = False
    ) -> None:
        self.show_gridlines = show_gridlines
        self.file_path = file_path
        self.file_template_path = file_template_path
        self.save_path = save_path
        self.adictional_rules_to_apply = adictional_rules_to_apply
        self.datas_validation_to_apply = datas_validation_to_apply
        
        self.set_work_book_file()
        self.set_work_book_template()
        print("files readed")
        self.save_file_informations()
        self.save_template_informations()
        print("default config aplied")

    def apply_adictional_rules(self):
        """
            ## Add Conditional Formatation for Each Sheet
            Pattern adictional Rule:
            {"<SHEET_NAME>": [
                    {"<COLUMN_NAME>": Rule}
                ]
            }
        """
        if self.adictional_rules_to_apply:
            for sheet_name, rules in self.adictional_rules_to_apply.items():
                rules:dict
                for rule_prop in rules:
                    rule_prop:dict
                    for column_name, rule in rule_prop.items():
                        self.wb_file[sheet_name].conditional_formatting.add(column_name, rule)
    
    def apply_datas_validation(self) -> None:
        """
        datas_validation_to_apply has to follow this pattern:
        self.datas_validation_to_apply = {
            "<SHEETNAME>":{
                "<CELLS_APPLY>": "<FORMULA>"
            }
        }
        """
        if self.datas_validation_to_apply:
            for sheetname, datas_validations in self.datas_validation_to_apply.items():
                datas_validations:dict
                for cells_apply, formula in datas_validations.items():        
                    dv = DataValidation(type="list", formula1=f"{formula}")
                    self.wb_file[sheetname].add_data_validation(dv)
                    dv.add(cells_apply)
                                  
    def set_work_book_file(self) -> None:
        self.wb_file = openpyxl.load_workbook(filename=self.file_path)
            
    def set_work_book_template(self) -> None:
        self.wb_template = openpyxl.load_workbook(filename=self.file_template_path)

    def save_file_informations(self) -> None:
        self.wb_file_columns = {
            sheetname: [re.sub(r"[^a-zA-Z]", "", cell_c.coordinate) for cell_c in self.wb_file[sheetname]["1"]]
            for sheetname in self.wb_file.sheetnames
        }
    
    def save_template_informations(self):
        self.wb_template_columns = {
            sheetname: [re.sub(r"[^a-zA-Z]", "", cell_c.coordinate) for cell_c in self.wb_template[sheetname]["1"]]
            for sheetname in self.wb_template.sheetnames
        }

    def prepare_work_book_file(self) -> None:...

    def get_template_prop(self, sheetname: str, row_number_id: str = "2") -> dict:
        template_pattern = self.wb_template[sheetname][row_number_id]
        cell_pattern = {}
        template_columns = {}
        for cell in template_pattern:
            cell:Cell
            column = re.sub(r"[^a-zA-Z]", "", cell.coordinate)
            cell_pattern[column] = {
                "font": Font(
                    name=cell.font.name,
                    size=cell.font.size,
                    bold=cell.font.bold,
                    italic=cell.font.italic,
                    color=cell.font.color,
                ),
                "alignment": Alignment(
                    horizontal=cell.alignment.horizontal,
                    vertical=cell.alignment.vertical,
                    wrap_text=cell.alignment.wrap_text,
                ),
                "border": Border(
                    left=Side(
                        border_style=cell.border.left.style,
                        color=cell.border.left.color,
                    ),
                    right=Side(
                        border_style=cell.border.right.style,
                        color=cell.border.right.color,
                    ),
                    top=Side(
                        border_style=cell.border.top.style, color=cell.border.top.color
                    ),
                    bottom=Side(
                        border_style=cell.border.bottom.style,
                        color=cell.border.bottom.color,
                    ),
                ),
                "fill": PatternFill(
                    fill_type=cell.fill.fill_type,
                    start_color=cell.fill.start_color,
                    end_color=cell.fill.end_color,
                ),
                "number_format": cell.number_format,
            }
            
            if cell.data_type == "f":
                cell_pattern[column]["formula"] = Translator(cell.value, origin=cell.coordinate)
            template_columns[column] = list(self.wb_template[sheetname][column])
        return {"sheetname": sheetname, "cells_styles": cell_pattern}

    def prepare_template_to_fill(self, wb_template: Worksheet, wb_file:Worksheet) -> None:
        if wb_file.max_column > wb_template.max_column:
            wb_template.insert_cols(wb_template.max_column, wb_file.max_column - wb_template.max_column)
        wb_template.insert_rows(3, wb_file.max_row + 5)  
        
    def fill_column_template(
        self,
        wb_file_sheet_filtered: Worksheet,
        wb_file_sheet_filtered_cells: list[Cell],
        column: str,
        template_props: SimpleNamespace,
    ) -> None:   
        """
            attributes from get_template_prop;
        """     
        for parameter, value in template_props.header_style["cells_styles"][column].items():
            wb_file_sheet_filtered[f"{column}1"]
            setattr(wb_file_sheet_filtered[f"{column}1"], parameter, value)

        for parameter, value in template_props.cells_style["cells_styles"][column].items():
            for cell_file in wb_file_sheet_filtered_cells[1:]:
                cell_file:Cell
                if parameter == "formula":
                    value:Translator
                    cell_file.value = value.translate_formula(cell_file.coordinate)
                else:
                    setattr(cell_file, parameter, value)
        
    def change_all_columns(self):
        columns_changes_instances = []
        sheetsnames = set(self.wb_file_columns) & set(self.wb_template_columns)
        
        template_props = {}
        print("começo")
        for sheetname in list(sheetsnames):
            self.prepare_template_to_fill(self.wb_template[sheetname], self.wb_file[sheetname])
            template_props[sheetname] = SimpleNamespace(
                cells_style =self.get_template_prop(sheetname), 
                header_style=self.get_template_prop(sheetname, 1)
            )
            
        for sheetname in sheetsnames:               
            if not self.show_gridlines:
                self.wb_file[sheetname].sheet_view.showGridLines = False
                
            for column in self.wb_file_columns[sheetname]:
                t = Thread(
                    target=self.fill_column_template,
                    args=(
                        self.wb_file[sheetname],
                        self.wb_file[sheetname][column],
                        column,
                        template_props[sheetname],
                    ),
                )
                
                columns_changes_instances.append(t)
                t.start()

            for t in columns_changes_instances:
                t.join()
                
        print("fim")
        self.apply_adictional_rules()
        self.apply_datas_validation()
        self.wb_file.save(self.save_path)
        

if __name__ == "__main__":
    inicio = time.time()

    t = StyleTemplatePandas(
        r"C:\Users\u1280820\MMC\Marsh Brasil Data Analytics - Python\NLOC autm\resultado_macro_um.xlsx",
        r"C:\Users\u1280820\MMC\Marsh Brasil Data Analytics - Python\NLOC autm\src\templates\template_macro_um.xlsx",
        "teste.xlsx",
    )
    t.change_all_columns()

    fim = time.time()
    tempo_de_execucao = fim - inicio

    print(f"Tempo de Execução: {tempo_de_execucao:.6f} segundos")
    